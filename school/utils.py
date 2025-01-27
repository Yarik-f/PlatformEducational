import openpyxl
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse

from .models import Schedule, Teacher, Class, Subject


def get_key_by_value(choices, value):
    for key, val in choices:
        if val == value:
            return key


def import_schedule_from_excel_with_update(file):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                lesson_number = row[0]
                school_class_name = row[1]
                day_of_week = row[2]
                subject_name = row[3]
                office = row[4]
                teacher_full_name = row[5]
                start_time = row[6]
                end_time = row[7]

                valid_days = [day[1] for day in Schedule.DAY_OF_WEEK_CHOICES]
                if day_of_week not in valid_days:
                    return False, f"Некорректный день недели: {day_of_week}"
                else:
                    day = get_key_by_value(Schedule.DAY_OF_WEEK_CHOICES, day_of_week)

                school_class, _ = Class.objects.get_or_create(name=school_class_name.strip())

                subject, _ = Subject.objects.get_or_create(name=subject_name.strip())

                teacher = None
                if teacher_full_name:
                    fio = teacher_full_name.split()
                    last_name = fio[0] if len(fio) > 0 else ""
                    first_name = fio[1] if len(fio) > 1 else ""
                    patronymic = fio[2] if len(fio) > 2 else ""
                    teacher = Teacher.objects.filter(
                        last_name=last_name.strip(),
                        first_name=first_name.strip(),
                        patronymic=patronymic.strip()
                    ).first()

                existing_schedule = Schedule.objects.filter(
                    lesson_number=lesson_number,
                    school_class=school_class,
                    day_of_week=day
                ).first()

                if existing_schedule:
                    if (
                        existing_schedule.subject != subject or
                        existing_schedule.office != office or
                        existing_schedule.teacher != teacher or
                        existing_schedule.start_time != start_time or
                        existing_schedule.end_time != end_time
                    ):
                        existing_schedule.subject = subject
                        existing_schedule.office = office
                        existing_schedule.teacher = teacher
                        existing_schedule.start_time = start_time
                        existing_schedule.end_time = end_time
                        existing_schedule.save()
                else:
                    Schedule.objects.create(
                        lesson_number=lesson_number,
                        school_class=school_class,
                        day_of_week=day,
                        subject=subject,
                        office=office,
                        teacher=teacher,
                        start_time=start_time,
                        end_time=end_time
                    )

        return True, "Расписание успешно импортировано!"
    except Exception as e:
        return False, f"Ошибка при обработке файла: {e}"


def export_schedule_to_excel(classroom=None):
    workbook = Workbook()
    sheet = workbook.active
    if classroom:
        sheet.title = f"Расписание_{classroom}"
    else:
        sheet.title = "Расписание"

    headers = ['Номер урока', 'Класс', 'День недели', 'Предмет', 'Кабинет', 'Учитель', 'Начало', 'Конец']
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    if classroom:
        schedules = Schedule.objects.filter(school_class=classroom)
    else:
        schedules = Schedule.objects.all()

    for row_num, schedule in enumerate(schedules, start=2):
        sheet[f"A{row_num}"] = schedule.lesson_number
        sheet[f"B{row_num}"] = schedule.school_class.name
        sheet[f"C{row_num}"] = schedule.get_day_of_week_display()
        sheet[f"D{row_num}"] = schedule.subject.name
        sheet[f"E{row_num}"] = schedule.office
        sheet[f"F{row_num}"] = f'{str(schedule.teacher.last_name)} {str(schedule.teacher.first_name)} {str(schedule.teacher.patronymic)}' if schedule.teacher else ""
        sheet[f"G{row_num}"] = schedule.start_time.strftime('%H:%M')
        sheet[f"H{row_num}"] = schedule.end_time.strftime('%H:%M')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="schedule.xlsx"'
    return response
def export_schedule_classroom_to_excel(classroom):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Расписание"

    headers = ['Номер урока', 'Класс', 'День недели', 'Предмет', 'Кабинет', 'Учитель', 'Начало', 'Конец']
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet[f"{column_letter}1"] = header

    schedules = Schedule.objects.filter(school_class=classroom)
    for row_num, schedule in enumerate(schedules, start=2):
        sheet[f"A{row_num}"] = schedule.lesson_number
        sheet[f"B{row_num}"] = schedule.school_class.name
        sheet[f"C{row_num}"] = schedule.get_day_of_week_display()
        sheet[f"D{row_num}"] = schedule.subject.name
        sheet[f"E{row_num}"] = schedule.office
        sheet[f"F{row_num}"] = f'{str(schedule.teacher.last_name)} {str(schedule.teacher.first_name)} {str(schedule.teacher.patronymic)}' if schedule.teacher else ""
        sheet[f"G{row_num}"] = schedule.start_time.strftime('%H:%M')
        sheet[f"H{row_num}"] = schedule.end_time.strftime('%H:%M')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="schedule.xlsx"'
    return response
